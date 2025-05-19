import locale
from datetime import date, timedelta

import pandas as pd
import reportlab.rl_config
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from streamlit.connections import SQLConnection

locale.setlocale(locale.LC_ALL, "pt_BR.UTF-8")

engine: SQLConnection = st.connection(name="DB2", type=SQLConnection)


def load_evid(_data_pos: date) -> pd.DataFrame:
    return engine.query(
        sql="""
            SELECT
                CAST(CD_CLI_ACNT AS BIGINT) AS MCI,
                SUM(CAST(QT_TIT_INC_MM AS BIGINT)) AS QUANTIDADE
            FROM
                DB2AEB.PSC_TIT_MVTD
            WHERE
                CD_CLI_ACNT IN ('423566153', '100225800', '100186582', '903485186', '903721460') AND
                DT_PSC = :data_posterior AND
                CD_CLI_EMT = 903485186
            GROUP BY
                CD_CLI_ACNT,
                DT_PSC,
                CD_CLI_EMT
            ORDER BY
                CD_CLI_ACNT
        """,
        ttl=0,
        show_spinner=False,
        params=dict(data_posterior=_data_pos),
    )


st.subheader(":material/published_with_changes: Autorregulação BB")

reportlab.rl_config.warnOnMissingFontGlyphs = 0

pdfmetrics.registerFont(TTFont("Vera", "Vera.ttf"))
pdfmetrics.registerFont(TTFont("VeraBd", "VeraBd.ttf"))
pdfmetrics.registerFont(TTFont("VeraIt", "VeraIt.ttf"))
pdfmetrics.registerFont(TTFont("VeraBI", "VeraBI.ttf"))

with st.columns(3)[0]:
    st.markdown("##### Gerar Evidências")
    col1, col2 = st.columns(2)
    col1.number_input("**Mês:**", min_value=1, max_value=12, value=date.today().month, key="mes_ev")
    col2.number_input("**Ano:**", min_value=2000, max_value=2100, value=date.today().year, key="ano_ev")
    st.button("**Arquivo PDF**", key="btn_ev", type="primary", icon=":material/picture_as_pdf:")

    st.markdown("##### Gerar Base Acionária")
    col1, col2 = st.columns(2)
    col1.number_input("**Mês:**", min_value=1, max_value=12, value=date.today().month, key="mes_ac")
    col2.number_input("**Ano:**", min_value=2000, max_value=2100, value=date.today().year, key="ano_ac")
    st.file_uploader("**Escolha o arquivo 738 correspondente:**", key="up_base_aci")
    st.button("**Arquivo Excel**", key="btn_ac", type="primary", icon=":material/upload:")

    st.markdown("##### Gerar Base Autorregulação")
    col1, col2 = st.columns(2)
    col1.number_input("**Mês:**", min_value=1, max_value=12, value=date.today().month, key="mes_au")
    col2.number_input("**Ano:**", min_value=2000, max_value=2100, value=date.today().year, key="ano_au")
    st.file_uploader("**Escolha o arquivo 738 correspondente:**", key="up_base_738")
    st.file_uploader("**Escolha o arquivo que chegou no Siri:**", key="up_base_siri")
    st.button("**Arquivo Excel**", key="btn_au", type="primary", icon=":material/upload:")

if st.session_state["btn_ev"]:
    data_pos: date = date(
        st.session_state["ano_ev"] + 1 if st.session_state["mes_ev"] == 12 else st.session_state["ano_ev"],
        1 if st.session_state["mes_ev"] == 12 else st.session_state["mes_ev"] + 1,
        1
    )

    dfevid: pd.DataFrame = load_evid(data_pos)

    label_evidence: list[dict[str, str]] = [
        {"empresa": "Caixa de Previdência Funcis Banco do Brasil", "cnpj": "33.754.482/0001-24"},
        {"empresa": "Ministério da Economia", "cnpj": "00.394.460/0001-41"},
        {"empresa": "The Bank of New York ADR Department", "cnpj": "05.523.773/0001-76"},
        {"empresa": "Banco do Brasil", "cnpj": "00.000.000/0001-91"},
        {"empresa": "BB Gestão de Recursos Dist Títulos e Val", "cnpj": "30.822.936/0001-69"},
    ]

    if dfevid.empty:
        st.toast("###### Não consta nada na data estabelecida...", icon=":material/error:")
        st.stop()

    last_day: date = data_pos - timedelta(days=1)

    cnv: canvas.Canvas = canvas.Canvas(
        filename=f"static/escriturais/@deletar/Autorregulação BB - Evidências - {data_pos:%B de %Y}.pdf",
        pagesize=A4
    )
    cnv.drawImage(image="static/imagens/bb.jpg", x=40, y=780, width=300, height=38)
    cnv.setFont(psfontname="Vera", size=11)

    cnv.drawString(x=350, y=730, text=f"Rio de Janeiro, {date.today():%d/%m/%Y}")
    cnv.drawString(x=40, y=680, text="Prezados Senhores,"),

    cnv.drawString(x=40, y=650, text=f"Informamos evidenciações de posição e participação societária das principais")
    cnv.drawString(x=40, y=635, text=f"instituições com relação direta ao acionista controlador do Banco do Brasil,")
    cnv.drawString(x=40, y=620, text=f"referente ao mês de {last_day:%B de %Y}.")

    coord_y: int = 580

    for x in range(len(dfevid)):
        cnv.drawString(x=40, y=coord_y, text=f"Empresa: {label_evidence[x]['empresa']}")
        cnv.drawString(x=350, y=coord_y, text=f"CNPJ: {label_evidence[x]['cnpj']}")
        cnv.drawString(x=40, y=coord_y - 15, text=f"Ações Ordinárias: {dfevid['quantidade'].iloc[x]}")
        cnv.drawString(x=350, y=coord_y - 15,
                       text=f"Participação: {dfevid['quantidade'].iloc[x] / 2865417020:.7%}".replace(".", ","))

        coord_y -= 50

    cnv.setFont(psfontname="Vera", size=8)
    cnv.drawString(x=40, y=90, text="_" * 129)
    cnv.drawString(x=40, y=75, text="DIRETORIA OPERAÇÕES - DIOPE")
    cnv.drawString(x=40, y=60, text="Gerência Executiva de Negócios em Serviços Fiduciários - Gerência de "
                                    "Escrituração e Trustee")
    cnv.drawString(x=40, y=45, text="Avenida República do Chile, 330 - 9º andar - Torre Oeste - Centro - "
                                    "Rio de Janeiro RJ")
    cnv.drawString(x=40, y=30, text="Telefone: (21) 3808-3715")
    cnv.drawString(x=40, y=15, text="Ouvidoria BB - 0800 729 5678")

    cnv.save()

    st.toast("###### Arquivo PDF gerado com sucesso e enviado na pasta específica", icon=":material/check_circle:")

if st.session_state["btn_ac"]:
    data_pos: date = date(
        st.session_state["ano_ac"] + 1 if st.session_state["mes_ac"] == 12 else st.session_state["ano_ac"],
        1 if st.session_state["mes_ac"] == 12 else st.session_state["mes_ac"] + 1,
        1
    )

    if not st.session_state["up_base_aci"]:
        st.toast("###### Ainda não baixou o arquivo 738 correspondente...", icon=":material/warning:")
        st.stop()

    with st.spinner("**:material/hourglass: Preparando os dados, aguarde...**", show_time=True):
        dados: pd.DataFrame = pd.read_fwf(
            filepath_or_buffer=st.session_state["up_base_aci"],
            colspecs=[(0, 9), (9, 24), (39, 89), (99, 101), (137, 140), (345, 360),
                      (387, 402), (429, 444), (471, 486), (513, 528), (555, 570)],
            names=["MCI", "CPF", "NOME", "TP", "PAIS", "qtd_bb_1", "qtd_cblc_1",
                   "qtd_bb_25", "qtd_cblc_25", "qtd_bb_27", "qtd_cblc_27"],
            encoding="latin"
        )
    
        dados["TOTAL"] = (dados["qtd_bb_1"] + dados["qtd_bb_25"] + dados["qtd_bb_27"] + dados["qtd_cblc_1"] +
                          dados["qtd_cblc_25"] + dados["qtd_cblc_27"])
    
        dados = dados[["MCI", "CPF", "NOME", "TP", "PAIS", "TOTAL"]]
        dados = dados[2:-4]
        dados = dados[dados["MCI"].ne(205007939)]
    
        # variáveis do primeiro quadro
        inv_pf_bra = dados[dados.TP.eq("PF") & dados.PAIS.eq("BRA")].count()[0]
        inv_pj_bra = dados[dados.TP.eq("PJ") & dados.PAIS.eq("BRA")].count()[0]
        invest = dados[dados.PAIS.ne("BRA")].count()[0]
        ac_pf_bra = sum(dados.TOTAL[dados.TP.eq("PF") & dados.PAIS.eq("BRA")])
        ac_pj_bra = sum(dados.TOTAL[dados.TP.eq("PJ") & dados.PAIS.eq("BRA")])
        acest = sum(dados.TOTAL[dados.PAIS.ne("BRA")])
    
        # variáveis do segundo quadro
        inv1a11 = dados[dados.TOTAL.ge(1) & dados.TOTAL.le(11)].count()[0]
        inv12a50 = dados[dados.TOTAL.ge(12) & dados.TOTAL.le(50)].count()[0]
        inv51a100 = dados[dados.TOTAL.ge(51) & dados.TOTAL.le(100)].count()[0]
        inv101a1000 = dados[dados.TOTAL.ge(101) & dados.TOTAL.le(1000)].count()[0]
        inv_mais_que_1000 = dados[dados.TOTAL.gt(1000)].count()[0]
    
        ac1a11 = sum(dados.TOTAL[dados.TOTAL.ge(1) & dados.TOTAL.le(11)])
        ac12a50 = sum(dados.TOTAL[dados.TOTAL.ge(12) & dados.TOTAL.le(50)])
        ac51a100 = sum(dados.TOTAL[dados.TOTAL.ge(51) & dados.TOTAL.le(100)])
        ac101a1000 = sum(dados.TOTAL[dados.TOTAL.ge(101) & dados.TOTAL.le(1000)])
        ac_mais_que_1000 = sum(dados.TOTAL[dados.TOTAL.gt(1000)])
    
    
        def arquivo_base_acionaria() -> None:
            # Create a Pandas Excel writer using XlsxWriter as the engine.
            writer: pd.ExcelWriter = pd.ExcelWriter(
                path=f"static/escriturais/@deletar/Base Acionária_-_DIOPE-GEFID_-_{data_pos:%B de %Y}.xlsx",
                engine="xlsxwriter"
            )
    
            # definindo abas
            workbook = writer.book
            worksheet_qr = workbook.add_worksheet("QuadroResumo")
            worksheet1 = workbook.add_worksheet("1")
            worksheet2 = workbook.add_worksheet("2")
    
            # exportando base
            dados[:int(1e6)].to_excel(writer, index=False, sheet_name="1", header=False, startrow=2)
            dados[int(1e6):].to_excel(writer, index=False, sheet_name="2", header=False, startrow=2)
    
            # definindo formatos
            menu_format = workbook.add_format(dict(align="center", bold=True, bg_color="#9BC2E6", font_size=16))
            text_format = workbook.add_format(dict(align="center", bold=False, font_size=16))
            qtd_format = workbook.add_format(dict(num_format="#,##0", align="center", font_size=16))
            number_format = workbook.add_format(dict(num_format="0"))
            titulo = workbook.add_format(dict(bold=True, align="center", bg_color="#025AA5", font_size=22,
                                              font_color="white"))
    
            # conteúdo da aba quadro resumo
            worksheet_qr.set_column(1, 3, 50)
            worksheet_qr.merge_range("B2:D2", f"Autorregulação Banco do Brasil - {data_pos:%B de %Y}", titulo)
    
            worksheet_qr.write("B4", "Segmento", menu_format)
            worksheet_qr.write("B5", "Pessoas Físicas no País", text_format)
            worksheet_qr.write("B6", "Pessoas Jurídicas no País", text_format)
            worksheet_qr.write("B7", "Capital Estrangeiro (PF+PF+ADR)", text_format)
            worksheet_qr.write("B8", "Totais", menu_format)
    
            worksheet_qr.write("C4", "Acionistas", menu_format)
            worksheet_qr.write("C5", inv_pf_bra, qtd_format)
            worksheet_qr.write("C6", inv_pj_bra, qtd_format)
            worksheet_qr.write("C7", invest, qtd_format)
            worksheet_qr.write("C8", inv_pf_bra + inv_pj_bra + invest, menu_format)
    
            worksheet_qr.write("D4", "Ações", menu_format)
            worksheet_qr.write("D5", ac_pf_bra, qtd_format)
            worksheet_qr.write("D6", ac_pj_bra, qtd_format)
            worksheet_qr.write("D7", acest, qtd_format)
            worksheet_qr.write("D8", ac_pf_bra + ac_pj_bra + acest, menu_format)
    
            worksheet_qr.write("B10", "Ações por Faixa", menu_format)
            worksheet_qr.write("B11", "1 a 11", text_format)
            worksheet_qr.write("B12", "12 a 50", text_format)
            worksheet_qr.write("B13", "51 a 100", text_format)
            worksheet_qr.write("B14", "101 a 1000", text_format)
            worksheet_qr.write("B15", "> 1000", text_format)
            worksheet_qr.write("B16", "Totais", menu_format)
    
            worksheet_qr.write("C10", "Qtd. Acionistas / Faixa", menu_format)
            worksheet_qr.write("C11", inv1a11, qtd_format)
            worksheet_qr.write("C12", inv12a50, qtd_format)
            worksheet_qr.write("C13", inv51a100, qtd_format)
            worksheet_qr.write("C14", inv101a1000, qtd_format)
            worksheet_qr.write("C15", inv_mais_que_1000, qtd_format)
            worksheet_qr.write("C16", inv1a11 + inv12a50 + inv51a100 + inv101a1000 + inv_mais_que_1000, menu_format)
    
            worksheet_qr.write("D10", "Qtd Total de Ações", menu_format)
            worksheet_qr.write("D11", ac1a11, qtd_format)
            worksheet_qr.write("D12", ac12a50, qtd_format)
            worksheet_qr.write("D13", ac51a100, qtd_format)
            worksheet_qr.write("D14", ac101a1000, qtd_format)
            worksheet_qr.write("D15", ac_mais_que_1000, qtd_format)
            worksheet_qr.write("D16", ac1a11 + ac12a50 + ac51a100 + ac101a1000 + ac_mais_que_1000, menu_format)
    
            worksheet_qr.outline_settings(False, False, False, True)
    
            # conteúdo da primeira parte da base
            worksheet1.set_column(0, 0, 12, number_format)
            worksheet1.set_column(1, 1, 16, number_format)
            worksheet1.set_column(2, 2, 40)
            worksheet1.set_column(3, 3, 6)
            worksheet1.set_column(4, 4, 6)
            worksheet1.set_column(5, 5, 16, number_format)
    
            worksheet1.write("A2", "MCI", menu_format)
            worksheet1.write("B2", "CPF/CNPJ", menu_format)
            worksheet1.write("C2", "Nome", menu_format)
            worksheet1.write("D2", "TP", menu_format)
            worksheet1.write("E2", "Pais", menu_format)
            worksheet1.write("F2", "Total", menu_format)
    
            # conteúdo da segunda parte da base
            worksheet2.set_column(0, 0, 12, number_format)
            worksheet2.set_column(1, 1, 16, number_format)
            worksheet2.set_column(2, 2, 40)
            worksheet2.set_column(3, 3, 6)
            worksheet2.set_column(4, 4, 6)
            worksheet2.set_column(5, 5, 16, number_format)
    
            worksheet2.write("A2", "MCI", menu_format)
            worksheet2.write("B2", "CPF/CNPJ", menu_format)
            worksheet2.write("C2", "Nome", menu_format)
            worksheet2.write("D2", "TP", menu_format)
            worksheet2.write("E2", "Pais", menu_format)
            worksheet2.write("F2", "Total", menu_format)
    
            # fecha arquivo
            workbook.close()
            writer.close()
    
    
        arquivo_base_acionaria()
    
        st.toast("###### Geração de Excel feita com sucesso!", icon=":material/check_circle:")

if st.session_state["btn_au"]:
    data_pos = date(
        st.session_state["ano_au"] + 1 if st.session_state["mes_au"] == 12 else st.session_state["ano_au"],
        1 if st.session_state["mes_au"] == 12 else st.session_state["mes_au"] + 1,
        1
    ) - timedelta(days=1)

    if not all([st.session_state["up_base_738"], st.session_state["up_base_siri"]]):
        st.toast("###### Não subiu os 2 arquivos exigidos...", icon=":material/warning:")
        st.stop()

    with st.spinner("**:material/hourglass: Preparando os dados, aguarde...**", show_time=True):
        dados: pd.DataFrame = pd.read_fwf(
            filepath_or_buffer=st.session_state.up_base_738,
            colspecs=[(0, 9), (9, 24), (39, 89), (99, 101), (137, 140), (345, 360),
                      (387, 402), (429, 444), (471, 486), (513, 528), (555, 570)],
            names=["MCI", "CPF", "NOME", "TP", "PAIS", "qtd_bb_1", "qtd_cblc_1",
                   "qtd_bb_25", "qtd_cblc_25", "qtd_bb_27", "qtd_cblc_27"],
            encoding="latin"
        )
    
        dados["TOTAL"] = (dados["qtd_bb_1"] + dados["qtd_bb_25"] + dados["qtd_bb_27"] + dados["qtd_cblc_1"] +
                          dados["qtd_cblc_25"] + dados["qtd_cblc_27"])
    
        dados = dados[["MCI", "CPF", "NOME", "TP", "PAIS", "TOTAL"]]
        dados = dados[2:-4]
        dados = dados[dados["MCI"].ne(205007939)]
    
    
        def arquivo_siri() -> None:
            # carrega arquivo Base Acionária
            base_acionaria = load_workbook(st.session_state.up_base_siri)
    
            def aba_base() -> None:
                mes_anterior: date = data_pos - timedelta(days=31)
    
                # escolhe a aba do mês anterior ao mês da autorregulação
                sheet1 = base_acionaria[f"Base Acionária - {mes_anterior:%B de %Y}"]
    
                # deleta colunas dos mes -2
                sheet1.delete_cols(idx=3, amount=3)
                sheet1.delete_cols(idx=6, amount=1)
    
                # copia colunas do mes passado para servir de referência
                for cell in sheet1["C:C"]:
                    sheet1.cell(row=cell.row, column=6, value=cell.value)
                for cell in sheet1["D:D"]:
                    sheet1.cell(row=cell.row, column=7, value=cell.value)
                for cell in sheet1["E:E"]:
                    sheet1.cell(row=cell.row, column=8, value=cell.value)
    
                # nomeia o título das colunas do mês de autorregulação
                sheet1["F1"] = data_pos.strftime("%d de %B de %Y")
    
                # corrigir / desnevolver
                mineco = sum(dados.TOTAL[dados.MCI.eq(100225800)])
                tesouro_bb = sum(dados.TOTAL[dados.MCI.eq(903485186)])
                tesouro_bb_asset = sum(dados.TOTAL[dados.MCI.eq(903721460)])
                ca = cd = de = 0
                previ = sum(dados.TOTAL[dados.MCI.eq(100186582)])
                outros_conselhos = 0
    
                # variáveis do primeiro quadro
                inv_pf_bra = dados[dados.TP.eq("PF") & dados.PAIS.eq("BRA")].count()[0]
                inv_pj_bra = dados[dados.TP.eq("PJ") & dados.PAIS.eq("BRA")].count()[0]
                invest = dados[dados.PAIS.ne("BRA")].count()[0]
    
                ac_pf_bra = sum(dados.TOTAL[dados.TP.eq("PF") & dados.PAIS.eq("BRA")])
                ac_pj_bra = sum(dados.TOTAL[dados.TP.eq("PJ") & dados.PAIS.eq("BRA")])
                acest = sum(dados.TOTAL[dados.PAIS.ne("BRA")])
    
                # variáveis do segundo quadro
                inv1a11 = dados[dados.TOTAL.ge(1) & dados.TOTAL.le(11)].count()[0]
                inv12a50 = dados[dados.TOTAL.ge(12) & dados.TOTAL.le(50)].count()[0]
                inv51a100 = dados[dados.TOTAL.ge(51) & dados.TOTAL.le(100)].count()[0]
                inv101a1000 = dados[dados.TOTAL.ge(101) & dados.TOTAL.le(1000)].count()[0]
                inv_mais_que_1000 = dados[dados.TOTAL.gt(1000)].count()[0]
    
                ac1a11 = sum(dados.TOTAL[dados.TOTAL.ge(1) & dados.TOTAL.le(11)])
                ac12a50 = sum(dados.TOTAL[dados.TOTAL.ge(12) & dados.TOTAL.le(50)])
                ac51a100 = sum(dados.TOTAL[dados.TOTAL.ge(51) & dados.TOTAL.le(100)])
                ac101a1000 = sum(dados.TOTAL[dados.TOTAL.ge(101) & dados.TOTAL.le(1000)])
                ac_mais_que_1000 = sum(dados.TOTAL[dados.TOTAL.gt(1000)])
    
                # inclusão das variáveis no arquivo
                sheet1["G4"] = mineco
                sheet1["G7"] = tesouro_bb
                sheet1["G8"] = tesouro_bb_asset
                sheet1["G12"] = previ
    
                sheet1["G17"] = ac_pf_bra
                sheet1["G18"] = ac_pj_bra
                sheet1["G19"] = acest
    
                sheet1["G24"] = ca + cd + de
                sheet1["G25"] = outros_conselhos
    
                sheet1["G52"] = ac1a11
                sheet1["G53"] = ac12a50
                sheet1["G54"] = ac51a100
                sheet1["G55"] = ac101a1000
                sheet1["G56"] = ac_mais_que_1000
    
                sheet1["F17"] = inv_pf_bra
                sheet1["F18"] = inv_pj_bra
                sheet1["F19"] = invest
    
                sheet1["F52"] = inv1a11
                sheet1["F53"] = inv12a50
                sheet1["F54"] = inv51a100
                sheet1["F55"] = inv101a1000
                sheet1["F56"] = inv_mais_que_1000
    
                # funções de formatos
                def set_border(ws, cell_range) -> None:
                    thin = Side(border_style="thin", color="000000")
                    for row in ws[cell_range]:
                        for _cell in row:
                            _cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
                def fonte_padrao(ws, cell_range) -> None:
                    f1 = Font(name="Arial", size=11, color="00000000")
                    ad = Alignment(horizontal="right")
                    for row in ws[cell_range]:
                        for _cell in row:
                            _cell.font = f1
                            _cell.alignment = ad
    
                def tit_yellow(ws, cell_range) -> None:
                    y1 = PatternFill(fill_type="solid", start_color="00FFFF00", end_color="00FFFF00")
                    a1 = Alignment(horizontal="center")
                    fb = Font(name="Arial", size=11, bold=True)
                    for row in ws[cell_range]:
                        for _cell in row:
                            _cell.fill = y1
                            _cell.font = fb
                            _cell.alignment = a1
    
                def sub_cinza(ws, cell_range) -> None:
                    c1 = PatternFill(fill_type="solid", start_color="00969696", end_color="00969696")
                    fb = Font(name="Arial", size=11, bold=True)
                    for row in ws[cell_range]:
                        for _cell in row:
                            _cell.fill = c1
                            _cell.font = fb
    
                def tit_blue(ws, cell_range) -> None:
                    b1 = PatternFill(fill_type="solid", start_color="0099CCFF", end_color="0099CCFF")
                    fb = Font(name="Arial", size=11, bold=True)
                    for row in ws[cell_range]:
                        for _cell in row:
                            _cell.fill = b1
                            _cell.font = fb
    
                def coluna_i(ws, cell_range) -> None:
                    for row in ws[cell_range]:
                        for _cell in row:
                            _cell.border = Border(left=Side(border_style="medium", color="000000"))
                            _cell = f"= H{cell.row}-E{cell.row}"
    
                # executando funções
                set_border(sheet1, "F1:H57")
                fonte_padrao(sheet1, "F1:H57")
                tit_yellow(sheet1, "F1:f3")
                sub_cinza(sheet1, "F2:h3")
                sub_cinza(sheet1, "F5:h6")
                sub_cinza(sheet1, "F11:h11")
                sub_cinza(sheet1, "F13:h13")
                sub_cinza(sheet1, "F16:h16")
                sub_cinza(sheet1, "F20:h20")
                sub_cinza(sheet1, "F42:h42")
                sub_cinza(sheet1, "F44:h45")
                sub_cinza(sheet1, "F48:h48")
                sub_cinza(sheet1, "F51:h51")
                tit_blue(sheet1, "F21:h21")
                tit_blue(sheet1, "F23:h23")
                tit_blue(sheet1, "F27:h27")
                tit_blue(sheet1, "F41:h41")
    
                sheet1["f23"].border = Border(top=Side(border_style="medium", color="000000"))
                sheet1["g23"].border = Border(top=Side(border_style="medium", color="000000"))
                sheet1["h23"].border = Border(top=Side(border_style="medium", color="000000"))
    
                coluna_i(sheet1, "i1:i57")
    
                sheet1["f57"].border = Border(bottom=Side(border_style="medium", color="000000"),
                                              left=Side(border_style="medium", color="000000"),
                                              right=Side(border_style="thin", color="000000"))
                sheet1["g57"].border = Border(bottom=Side(border_style="medium", color="000000"),
                                              left=Side(border_style="thin", color="000000"),
                                              right=Side(border_style="thin", color="000000"))
                sheet1["h57"].border = Border(bottom=Side(border_style="medium", color="000000"),
                                              left=Side(border_style="thin", color="000000"))
    
                sheet1["f57"].font = Font(name="Arial", size=11, bold=True)
                sheet1["g57"].font = Font(name="Arial", size=11, bold=True)
                sheet1["h57"].font = Font(name="Arial", size=11, bold=True)
    
                sheet1.title = f"Base Acionária - {data_pos:%B} de {data_pos:%Y}"
    
            def aba_diretoria() -> None:
                # carrega aba
                sheet2 = base_acionaria["Membros da Diretoria"]
    
                dados["CPF"] = dados["CPF"].astype("int64")
    
                print(dados.head())
    
                for row in sheet2["C4:C210"]:
                    for cell in row:
                        if cell.value != "":
                            posicao = dados.query(f"CPF == {cell.value}")["TOTAL"].reset_index()
                            if posicao.shape[0] > 0:
                                sheet2[f"E{cell.row}"] = posicao["TOTAL"][0]
                            else:
                                sheet2[f"E{cell.row}"] = "-"
                        else:
                            continue
    
                base_acionaria.save(filename=f"static/escriturais/@deletar/Autorregulação_-_DIOPE-GEFID_-_"
                                             f"{data_pos:%B de %Y}.xlsx")
    
            aba_base()
            aba_diretoria()
    
    
        arquivo_siri()
    
        st.toast("###### Geração de Excel feita com sucesso!", icon=":material/check_circle:")
